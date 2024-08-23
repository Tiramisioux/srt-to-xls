
import os
from flask import Flask, render_template, request, send_file, redirect, url_for
import pandas as pd
from openpyxl import load_workbook
from datetime import datetime
import csv

app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = 'uploads'

# Function to convert SRT time to HH:MM:SS:ff
def srt_time_to_frames(time_str, fps=25):
    time_obj = datetime.strptime(time_str, "%H:%M:%S,%f")
    total_seconds = time_obj.hour * 3600 + time_obj.minute * 60 + time_obj.second + time_obj.microsecond / 1_000_000
    frames = int(total_seconds * fps) % fps
    timecode = time_obj.strftime(f"%H:%M:%S:{frames:02d}")
    return timecode

# Convert SRT to CSV
def convert_srt_to_csv(srt_file, csv_file, fps=25):
    with open(srt_file, 'r', encoding='utf-8') as srt, open(csv_file, 'w', newline='', encoding='utf-8') as csv_out):
        writer = csv.writer(csv_out)
        writer.writerow(["ID", "Start Timecode", "End Timecode", "Subtitle Text"])
        
        subtitle_id = 1
        subtitle_text = []
        
        for line in srt:
            line = line.strip()
            if line.isdigit():
                if subtitle_text:
                    writer.writerow([subtitle_id, start_timecode, end_timecode, ' '.join(subtitle_text)])
                    subtitle_id += 1
                    subtitle_text = []
            elif '-->' in line:
                start_str, end_str = line.split(' --> ')
                start_timecode = srt_time_to_frames(start_str, fps)
                end_timecode = srt_time_to_frames(end_str, fps)
            elif line:
                subtitle_text.append(line)
        
        if subtitle_text:
            writer.writerow([subtitle_id, start_timecode, end_timecode, ' '.join(subtitle_text)])

# Import CSV into Excel template and modify the canvas sheet
def import_csv_to_excel_template(csv_file, template_file, output_file):
    # Load the Excel template
    wb = load_workbook(template_file)
    
    # Load CSV data into a DataFrame
    df = pd.read_csv(csv_file)
    
    # Get the 'raw' sheet and clear existing data
    raw_sheet = wb['raw']
    for row in raw_sheet.iter_rows(min_row=2, max_row=raw_sheet.max_row, min_col=1, max_col=raw_sheet.max_column):
        for cell in row:
            cell.value = None
    
    # Append DataFrame rows to the 'raw' sheet starting from the second row
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), start=2):
        for c_idx, value in enumerate(row, start=1):
            raw_sheet.cell(row=r_idx, column=c_idx, value=value)

    # Modify the canvas sheet
    canvas_sheet = wb['canvas']
    
    # Remove intermediary timecodes in columns G and H
    for row in canvas_sheet.iter_rows(min_row=2, max_row=canvas_sheet.max_row, min_col=7, max_col=8):
        for cell in row:
            cell.value = None
    
    # Update columns B and C with VLOOKUP formulas
    for row in canvas_sheet.iter_rows(min_row=2, max_row=canvas_sheet.max_row, min_col=2, max_col=3):
        cell_b, cell_c = row
        cell_b.value = f"=VLOOKUP(A{cell_b.row}, raw!A:D, 2, FALSE)"
        cell_c.value = f"=VLOOKUP(A{cell_c.row}, raw!A:D, 3, FALSE)"

    # Save the new Excel file
    wb.save(output_file)

@app.route('/', methods=['GET', 'POST'])
def index():
    if request.method == 'POST':
        # Check if the post request has the file part
        if 'srt_file' not in request.files or 'template_file' not in request.files:
            return redirect(request.url)
        
        srt_file = request.files['srt_file']
        template_file = request.files['template_file']
        
        if srt_file.filename == '' or template_file.filename == '':
            return redirect(request.url)
        
        # Save the uploaded files
        srt_file_path = os.path.join(app.config['UPLOAD_FOLDER'], srt_file.filename)
        template_file_path = os.path.join(app.config['UPLOAD_FOLDER'], template_file.filename)
        srt_file.save(srt_file_path)
        template_file.save(template_file_path)

        # Process the SRT and template file
        csv_file_path = os.path.join(app.config['UPLOAD_FOLDER'], 'output.csv')
        output_file_path = os.path.join(app.config['UPLOAD_FOLDER'], 'output_with_subtitles.xlsx')

        convert_srt_to_csv(srt_file_path, csv_file_path)
        import_csv_to_excel_template(csv_file_path, template_file_path, output_file_path)

        return send_file(output_file_path, as_attachment=True, download_name='output_with_subtitles.xlsx')

    return render_template('index.html')

if __name__ == '__main__':
    app.run(debug=True)
